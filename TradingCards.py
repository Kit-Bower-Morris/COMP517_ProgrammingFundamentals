# Kit Bower-Morris
# Student Number: 201532917
# COMP517 - CA5 - Trading Cards

import openpyxl


class Card:

    """
    Card class is used by the Deck class when accessing individual cards.
    Does not need to be called upon outside of the Deck class.  
    """
  
    def __init__(self, theName, theType, theHP, theMoves, isShiny):
        """
        Initaliser for indiviual cards. Doesn't need to be called by user. Should be used by the Class:Deck functions.
        A card needs a Name, a Type, HP, List of Moves, and whether or not it is Shiny.
        """
        self.name = str(theName)
        self.type = theType
        self.HP = theHP
        self.moves = theMoves
        self.shiny = isShiny
          

    def __str__(self):
        """
        ___str___ function is used when a indiviual card is called to be printed. Should be used by the Class:Deck functions.

        """
        print("\nName: {self.name}".format(self=self))
        print("Type: {self.type}".format(self=self))
        print("Maximum HP: {self.HP}".format(self=self))
        print("Moves:")
        for i in self.moves:
            print(i)
        if(self.shiny == 0):
            return "Shiny Status: Not Shiny"
        else:     
            return "Shiny Status: Shiny!"
        

    def save(self):
        """
        Save function is called by the Save To File function in the Deck Class. It returns the card varibles as a list, so that it can be appended to a new save file. 
        The for loop here removes the moves from a list format and instead appends the individual varibles to the card list. 
        """
        card = [self.name, self.type, self.HP, self.shiny]
        for i in self.moves:
            for x in i:
                card.append(x)
        
        return card


class Deck:

    #deckOrder is used as a way of registering the index of the deck, as well as creating a robust signature for each card. 
    deckOrder = []


    def __init__(self):
        """
        Contains a list which contains the cards in the deck. 
        """
        self.deck = []


    def getDeck(self):
        """
        Function Get Deck returns the self.deck varible. A list of the cards in the deck. 
        """
        return self.deck


    def inputFromFile(self, fileName):
        """
        Input From File function reads an spreadsheet file containing card varibles as rows. 
        File should end with '.xlsx'.
        File should be in the Format of Name, Type, HP, Shiny, Move 1, Damage, Move 2, Damage, Move 3, Damage, Move 4, Damage, Move 5, Damage.
        This function then compiles this as a single card varible. 
        If the row is missing data or data is in the wrong format the card will not be registered. In this case the function should print the reason. 
        """
        
        self.fileName = str(fileName)

        book = openpyxl.load_workbook(self.fileName)
        sheet = book.active

        for row in sheet.iter_rows(min_row = 2):            

            card = []
            for cell in row:             
                card.append(cell.value)
            
            if(len(card) < 6):
                continue
            
            else:

                if(card[0] is None):
                    continue

                else:
                    Name = row[0].value
                    Name = str(Name)                

                    Type = row[1].value
                    if(Type != "Magi" and Type != "Water" and Type != "Fire" and Type != "Earth" and Type != "Air" and Type != "Astral"):
                        print("Card Type Invalid.", Name, " not registered.")
                        print("Valid Card Types are: 'Magi', 'Water', 'Fire', 'Earth', 'Air' or 'Astral'. Ensure first letter is Uppercase.")
                        pass
                    else:
                        hp = row[2].value
                        try:
                            float(hp)
                        except ValueError:
                            print("HP entered not valid. Please ensure all HP's are integers.", Name, "not registered.")
                            pass
                        except TypeError:
                            print("HP entered not valid. Please ensure all HP's are integers.", Name, "not registered.")
                            pass
                        else:
                            if(type(hp) != int):
                                hp = int(hp)
                                print("HP should be entered as an integer.", Name,"'s HP has been rounded down to", hp)

                            shiny = row[3].value
                            if(shiny != 0 and shiny !=1):
                                print("Shiny status should be entered as either '0' meaning false or '1' meaning true.", Name,"'s shiny status has been set to '0'")
                                shiny = 0

                            length = len(card)
                            
                            moves = []
                            movesOut = []
                            for i in range(4, length, 2):
                                singleMove = []
                                if(row[i].value is None):
                                    pass
                                else:
                                    singleMove.append(row[i].value)
                                    if(row[i+1].value is None):
                                        singleMove.append(0)
                                        moves.append(singleMove)
                                    else:
                                        singleMove.append(row[i+1].value)
                                        moves.append(singleMove)

                            if((len(moves)) > 5):
                                print(Name, " has too many moves. Only the first five will be registered.")
                                del moves[5:]

                            for i in moves:
                                if(len(i) != 2):
                                    print("Move needs to be entered as a list containing name as a string at index [0] and damage as an int at index [1].")
                                    print(i, " removed as a move for", Name)
                                    movesOut.append(moves.index(i))
                                else:    
                                    if (isinstance(i[0], str) == False):
                                        i[0] = str(i[0])
                                    if (isinstance(i[1], int) == False):
                                        try:
                                            int(i[1])
                                        except ValueError:
                                            print("Data type for move ", i[0], " invalid. Damage amount for this move set to 0")
                                            i[1] = 0
                                        i[1] = int(i[1])
                            movesOut.reverse()
                            for i in movesOut:
                                del moves[i] 
                        
                            self.deck.append(Card(Name, Type, hp, moves, shiny))
                            
                            
                            numOfMoves = len(moves)
                            totalDamage = 0
                            for i in moves:
                                totalDamage += i[1]
                            avDamage = (totalDamage/numOfMoves)   

                            Deck.deckOrder.append([Name, Type, hp, avDamage, shiny])

        self.numOfCards = len(self.deck)


    def __str__(self):
        """
        Printing the individual varibles in self.deck calls the ___str___ function in the Card Class. This means that each card is printed. 
        This function is called by the View All Cards function. 
        Returns an empty string to fufil requirements for a ___str___ function. 
        """
        self.numOfCards = len(self.deck)
        for i in range(0, self.numOfCards):
            print(Deck.getDeck(self)[i])
        return ""


    def addCard(self, theCard):
        """
        Function that is used by the user to add a new card to the deck. 
        New card needs to be in the right format: [Name, Type, HP, [[Move1, Damage], [Move2, Damage],...]]
        Function makes sure all entered varibles of the new card are present and in the right format. 
        If a value is wrong the card will not be registered and the function should print the reason.
        """
        if(isinstance(theCard, list)):
            if(len(theCard) == 5):
                Name = theCard[0]
                Type = theCard[1]
                hp = theCard[2]
                moves = theCard[3]
                shiny = theCard[4]
            else:
                print("Card does not contain all elements, Card not added.")
                return
        else:
            print("Card should be input as a list.")
            print("List should be in the format of: [Name, Type, HP, [[Move1, Damage], [Move2, Damage],...]]") 
            return  

        Name = str(Name)                

        if(Type != "Magi" and Type != "Water" and Type != "Fire" and Type != "Earth" and Type != "Air" and Type != "Astral"):
            print("Card Type Invalid.", Name, " not registered.")
            print("Valid Card Types are: 'Magi', 'Water', 'Fire', 'Earth', 'Air' or 'Astral'. Ensure first letter is Uppercase.")
            return
        else:
            try:
                float(hp)
            except ValueError:
                print("HP entered not valid. Please ensure all HP's are integers.", Name, "not registered.")
                return
            except TypeError:
                print("HP entered not valid. Please ensure all HP's are integers.", Name, "not registered.")
                return
            else:
                if(type(hp) != int):
                    hp = int(hp)
                    print("HP should be entered as an integer.", Name,"'s HP has been rounded down to", hp)

                if(shiny != 0 and shiny !=1):
                    print("Shiny status should be entered as either '0' meaning false or '1' meaning true.", Name,"'s shiny status has been set to '0'")
                    shiny = 0

                if((isinstance(moves, list)) == False):
                    print("Move needs to be entered as a list of lists. Containing name as a string at index [0] and damage as an int at index [1] of each list.")
                    return
                else:
                    for i in moves:
                        if((isinstance(i, list)) == False):
                            print("Move needs to be entered as a list containing name as a string at index [0] and damage as an int at index [1].")
                            return

                if((len(moves)) > 5):
                    print(Name, " has too many moves. Only the first five will be registered.")
                    del moves[5:]

                movesOut = []
                for i in moves:
                    if(len(i) != 2):
                        print("Move needs to be entered as a list containing name as a string at index [0] and damage as an int at index [1].")
                        print(i, " removed as a move for", Name)
                        movesOut.append(moves.index(i))
                    else:    
                        if (isinstance(i[0], str) == False):
                            i[0] = str(i[0])
                        if (isinstance(i[1], int) == False):
                            try:
                                int(i[1])
                            except ValueError:
                                print("Data type for move ", i[0], " invalid. Damage amount for this move set to 0")
                                i[1] = 0
                            i[1] = int(i[1])
                movesOut.reverse()
                for i in movesOut:
                    del moves[i] 
            
                self.deck.append(Card(Name, Type, hp, moves, shiny))
                
                
                numOfMoves = len(moves)
                totalDamage = 0
                for i in moves:
                    totalDamage += i[1]
                avDamage = (totalDamage/numOfMoves)   

                Deck.deckOrder.append([Name, Type, hp, avDamage, shiny])     
    

    def rmCard(self, theCard):
        """
        Function is used to remove a card from the deck. 
        theCard needs to be in the format given by the getcards function. i.e. needs to be in the format of a varible in the deckOrder list. 
        This then removes the information about that card. 
        """
        try:
            position = Deck.deckOrder.index(theCard)
        except ValueError:
            print(theCard, "was not found in deck. No card has been removed.")
        else:
            position = Deck.deckOrder.index(theCard)
            del self.deck[position]
            del Deck.deckOrder[position]


    def getMostPowerful(self):
        """
        Function which returns the most powerful card. 
        deckOrder list already has the average damages for all the cards. This function searches for the first card with the highest average damage. 
        """
        mostPowerful = 0
        counter = 0
        for i in Deck.deckOrder:
            if(i[3] > mostPowerful):
                mostPowerful = i[3]
                posistion = counter
            counter += 1
        
        return Deck.getDeck(self)[posistion]
        

    def getAverageDamage(self):
        """
        Function finds average damage of the whole deck. 
        deckOrder list already has the average damages for all the cards. 
        Function pluses these together and divides by the number of cards in the deck.
        """
        totalAvDam = 0
        self.numOfCards = len(self.deck)
        for i in Deck.deckOrder:
            totalAvDam += i[3]
        totalAvDam = round((totalAvDam/self.numOfCards), 1)
        
        return totalAvDam


    def viewAllCards(self):
        """
        This function is calls the ___str___ function in the deck. This prints all the cards in the deck.   
        """
        Deck.__str__(self)


    def viewAllShinyCards(self):
        """
        Function that finds all the shiny cards in the deck and prints them to the screen. 
        If shiny is 1 then the card is shiny and is registered here. 
        """
        position = []
        for i in Deck.deckOrder:
            if(i[4] == 1):
                position.append(Deck.deckOrder.index(i))
        for i in position:
            print(Deck.getDeck(self)[i])


    def viewAllByType(self, theType):
        """
        Function prints all cards of a certain type. 
        Types are case sensitive and must be one of the 6 known types. 
        """
        self.type = theType
        if(self.type != "Magi" and self.type != "Water" and self.type != "Fire" and self.type != "Earth" and self.type != "Air" and self.type != "Astral"):
            print("Requested Card Type Invalid.")
            print("Valid Card Types are: 'Magi', 'Water', 'Fire', 'Earth', 'Air' or 'Astral'. Ensure first letter is Uppercase and is input as a string.")
        else:
            position = []
            for i in Deck.deckOrder:
                if(i[1] == self.type):
                    position.append(Deck.deckOrder.index(i))
            for i in position:
                print(Deck.getDeck(self)[i])


    def getCards(self):
        """
        This function returns the cards in the deck as a list. Deck Order is a list which contains the cards Name, Type, HP, Shiny and Average Damage. 
        Function can also be used to get an indiviual card by calling getCards()[i] where i is the index of the postion of the card in the deck. 
        """
        return Deck.deckOrder


    def saveToFile(self,fileName):
        """
        Function Save To File creates a new spreadsheet with the current cards in the deck. 
        File Name needs to have '.xlsx' at the end. If this is not the case this function adds it to it. 
        """
        fileName = str(fileName)
        if((fileName.endswith(".xlsx")) == False):
            fileName += ".xlsx"

        self.numOfCards = len(self.deck)
        log = []
        for i in Deck.getDeck(self):
            log.append(Card.save(i))
            
        newBook = openpyxl.Workbook()
        newSheet = newBook.active

        firstRow = ("Name", "Type", "HP", "Shiny", "Move Name 1", "Damage 1", "Move Name 2", "Damage 2", "Move Name 3", "Damage 3", "Move Name 4", "Damage 4", "Move Name 5", "Damage 5")
        newSheet.append(firstRow)
        
        for i in log:
            newSheet.append(i)
        
        newBook.save(fileName)


def main():
    """
    Main Function will only be called if file name is the same. 
    In this function find an example of a new card to be added to the deck. This shows the format in which it should be inserted. 
    
    theDeck = Deck()
    theDeck.inputFromFile("sampleDeck.xlsx")

    theDeck.addCard(["Name", "Water", 50, [["Move1", 50], ["Move2", 50]], 1])
    """

if __name__ == "__main__":
    main()